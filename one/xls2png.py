from win32com.client import Dispatch, DispatchEx
from PIL import ImageGrab
import pythoncom
import uuid


# screen_area——类似格式"A1:J10"
def excel_catch_screen(filename, sheetname, screen_area, img_name=False):
    """ 对excel的表格区域进行截图——用例：excel_catch_screen(ur"D:\Desktop\123.xlsx", "Sheet1", "A1:J10")"""
    pythoncom.CoInitialize()  # excel多线程相关

    excel = DispatchEx("Excel.Application")  # 启动excel
    excel.Visible = True  # 可视化
    excel.DisplayAlerts = False  # 是否显示警告
    wb = excel.Workbooks.Open(filename)  # 打开excel
    ws = wb.Sheets(sheetname)  # 选择sheet
    # ws.Range(screen_area).CopyPicture()  # 复制图片区域
    ws.Range(screen_area).CopyPicture()
    ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置

    name = str(uuid.uuid4())  # 重命名唯一值
    new_shape_name = name[:6]
    excel.Selection.ShapeRange.Name = new_shape_name  # 将刚刚选择的Shape重命名，避免与已有图片混淆

    ws.Shapes(new_shape_name).Copy()  # 选择图片
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    if not img_name:
        img_name = name + ".PNG"
    img.save("D:\\"+img_name)  # 保存图片
    wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
    excel.Quit()  # 退出excel
    pythoncom.CoUninitialize()





import openpyxl
def sheetname_get(filename):
    wb = openpyxl.load_workbook(filename)
    #功能1：打印工作表名称
    sheet_name = wb.sheetnames
    print(sheet_name)
    if len(sheet_name)==1:
        return sheet_name[0]
    else:
        return sheet_name[0]

if __name__ == '__main__':
    pass
    path = "D:\\1.xlsx"
    sheet_name = sheetname_get(path)
    sheet_name = sheet_name.replace('[','').replace(']','')
    excel_catch_screen(path, sheet_name, "A1:H240")

def sheetname_area(filename):
    wb = openpyxl.load_workbook(filename)
    #功能1：打印工作表名称
    sheet_name = wb.sheetnames
    sheet = sheet_name[0]